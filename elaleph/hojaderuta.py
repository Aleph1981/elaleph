# -*- coding: utf-8 -*-
"""
Created on Mon Feb  8 13:38:55 2021

@author: aleja
"""

from PyQt5 import QtSql, QtCore, QtGui, QtWidgets
from hojaderuta_ui import *
from bdstd import BdStd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border,Side, Alignment
from openpyxl.styles import colors
from openpyxl.cell import Cell
from bdstd import *
from datetime import timedelta, datetime
import datetime
from configctx import *

class HojaDeRuta(QtWidgets.QDialog, HojaDeRuta_Ui):
    
    def __init__(self):
        QtWidgets.QDialog.__init__(self)
        self.ui = HojaDeRuta_Ui()
        self.ui.setupUi(self)
        self.year = datetime.datetime.today().year
        for i in range (0,5):
            self.ui.comboAnyo.addItem(str(self.year-i))
        self.dic_meses={"Enero":"-01","Febrero":"-02","Marzo":"-03","Abril":"-04",\
                        "mayo":"-05","Junio":"-06","Julio":"-07","Agosto":"-08",\
                        "Septiembre":"-09","Octubre":"-10","Noviembre":"-11","Diciembre":"12"}
        
        self.ui.comboMes.activated.connect(self.load_tabla)
        self.ui.tableEventos.setColumnCount(4)
        self.ui.tableEventos.setHorizontalHeaderLabels(["ID","Nombre","Cliente","Fecha"])
        self.ui.tableEventos.setSelectionBehavior(QtWidgets.QTableWidget.SelectRows)
        self.ui.tableEventos.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.ui.tableEventos.setSelectionMode(QtWidgets.QTableWidget.SingleSelection)
        self.ui.tableEventos.verticalHeader().hide()
        self.ui.buttonCrear.clicked.connect(self.crear_hoja)
        
        ctx = ConfigCtx()
        self.carpeta = ctx.readvar("RUTAS", "hoja_de_ruta")
        
    def load_tabla(self):
        self.ui.tableEventos.setRowCount(0)
        self.year=self.ui.comboAnyo.currentText()
        self.mes=self.ui.comboMes.currentText()
        self.mes_num=self.dic_meses[self.mes]
        self.fecha=f"{self.year}{self.mes_num}"
        bd=BdStd()
        bd.runsql(f"""SELECT strftime('%d-%m-%Y', di.fecha),di.id_evento,ev.nombre,di.tarea,ev.cliente
                      FROM dias_evento as di, evento as ev
                      WHERE ev.id_evento=di.id_evento
                      AND di.fecha LIKE '{self.fecha}%' GROUP BY di.id_evento""")
        if bd.rows != None:
            for row in bd.rows:
                rowPosition = self.ui.tableEventos.rowCount()
                self.ui.tableEventos.insertRow(rowPosition)
                self.ui.tableEventos.setItem(rowPosition , 0, QtWidgets.QTableWidgetItem(row[1]))
                self.ui.tableEventos.setItem(rowPosition , 1, QtWidgets.QTableWidgetItem(row[2]))
                self.ui.tableEventos.setItem(rowPosition , 2, QtWidgets.QTableWidgetItem(row[3])) 
                self.ui.tableEventos.setItem(rowPosition , 3, QtWidgets.QTableWidgetItem(row[0])) 
                
    def crear_hoja(self):
        
        #try:
            redFill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
            blueFill= PatternFill(start_color="FF0075FF", end_color="FF0050FF", fill_type="solid")
            magentaFill= PatternFill(start_color="FFFF50FF", end_color="FFFF50FF", fill_type="solid")
            
            thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
            medium_border=Border(left=Side(style='medium'), 
                         right=Side(style='medium'), 
                         top=Side(style='medium'), 
                         bottom=Side(style='medium'))
            
            row=self.ui.tableEventos.currentRow()
            if row >= 0:
                self.id_evento=self.ui.tableEventos.item(row,0).text()
                print(self.id_evento)
            
            
            wb = load_workbook(filename = self.carpeta + "hojaderutabasevacia.xlsx")
            ws1 = wb.active
            bd = BdStd()
            
            
            ws1.title = "Hoja de ruta"
            
            bd.runsql(f"""SELECT ev.id_evento,ev.nombre,re.nombre,re.direccion||' '|| re.ciudad,cliente,
                      contacto_onsite,telefono_onsite,email_onsite, ev.notas, ma.nombre||''||ma.apellidos
                      FROM evento as ev,recintos as re, managers as ma 
                      WHERE ev.id_evento = '{self.id_evento}' AND re.id_recinto = ev.id_recinto AND
                      ma.id_manager = ev.id_manager;""")
            
            if bd.rows!=None:
                row=bd.rows[0]
                for i in range(0,8):
                    ws1[f"C{15+i}"]=row[i]
                    
                
                ws1["C24"]=row[8]
                ws1["C27"]=row[9]
                
            bd.runsql(f"""SELECT id_evento, strftime('%d-%m /', MIN(fecha))
                      ||''||strftime('%d-%m-%Y', MAX(fecha)) FROM dias_evento
                      WHERE id_evento='{self.id_evento}';""")
                      
            if bd.rows!=None:
                ws1["C23"]=bd.rows[0][1]
                
            bd.runsql(f"SELECT fecha, hora,tarea FROM dias_evento WHERE id_evento='{self.id_evento}';")
            print(bd.rows)
            
            xlrow=30
            if bd.rows != None:
                for fecha in bd.rows:
                #--------------------celdas para las fechas-----------------------------------------------
                
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow+1, end_column=7)
                    
                    ws1[f"A{xlrow}"]=bd.gira_fecha(fecha[0])+" "+fecha[1]+" "+fecha[2]
                    ws1[f"A{xlrow}"].fill = redFill
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=14, bold=True)
                    ws1[f"A{xlrow}"].border = medium_border
                    
                    xlrow+=3
                    
                #-------------------cabecera para el personal-----------------------------------------------
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=3)
                    ws1.merge_cells(start_row=xlrow, start_column=4, end_row=xlrow, end_column=5)
                    ws1.merge_cells(start_row=xlrow, start_column=6, end_row=xlrow, end_column=7)
                    ws1[f"A{xlrow}"]="Nombre"
                    ws1[f"A{xlrow}"].fill = redFill
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"A{xlrow}"].border = thin_border
                    ws1[f"D{xlrow}"]="Teléfono"
                    ws1[f"D{xlrow}"].fill = redFill
                    ws1[f"D{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"D{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"D{xlrow}"].border = thin_border
                    ws1[f"F{xlrow}"]="Cargo"
                    ws1[f"F{xlrow}"].fill = redFill
                    ws1[f"F{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"F{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"F{xlrow}"].border = thin_border
                    
                    xlrow+=1
                    
                 #-----------Select y coloca el personal-------------------------------------------------------
                 
                    bdper=BdStd()
                    bdper.runsql(f"""SELECT fecha,per.nombre||""||per.apellidos,per.telefono,ca.nombre 
                                 FROM personal_evento pe,personal as per,cargos as ca WHERE pe.id_evento='{self.id_evento}'
                                 AND per.id_personal=pe.id_personal AND ca.id_cargo=pe.id_cargo AND fecha='{fecha[0]}' AND hora='{fecha[1]}'
                                 ORDER BY ca.id_cargo;""")
                    
                    if bdper.rows != None:
                        for persona in bdper.rows:
                            
                            ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=3)
                            ws1.merge_cells(start_row=xlrow, start_column=4, end_row=xlrow, end_column=5)
                            ws1.merge_cells(start_row=xlrow, start_column=6, end_row=xlrow, end_column=7)
                            ws1[f"A{xlrow}"]=persona[1]
                            ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                            ws1[f"A{xlrow}"].font= Font(name='Arial',size=8)
                            ws1[f"A{xlrow}"].border = thin_border
                            ws1[f"D{xlrow}"]=persona[2]
                            ws1[f"D{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            ws1[f"D{xlrow}"].font= Font(name='Arial',size=8)
                            ws1[f"D{xlrow}"].border = thin_border
                            ws1[f"F{xlrow}"]=persona[3]
                            ws1[f"F{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            ws1[f"F{xlrow}"].font= Font(name='Arial',size=8)
                            ws1[f"F{xlrow}"].border = thin_border
                            xlrow+=1
                    xlrow+=1
            xlrow+=1
            
            #---------------TRANSPORTES---------------------------------------
            bdtrans=BdStd()
            bdtrans.runsql(f"""SELECT te.id_proveedor, te.conductor,te.conductor_telefono,te.conductor_email,
                           te.tipo_vehiculo,te.matricula,te.notas,te.recogida_lugar,te.recogida_localidad, te.recogida_direc,
                           te.recogida_indicaciones,te.recogida_fecha, te.recogida_hora, 
                           te.entrega_lugar,te.entrega_localidad,te.entrega_direc,te.entrega_indicaciones,
                           te.entrega_fecha,te.entrega_hora FROM transporte_evento as te 
                           WHERE id_evento = '{self.id_evento}' ORDER BY recogida_fecha, recogida_hora""")
            
            if len(bdtrans.rows) == 1:
                
                data=bdtrans.rows[0]
                ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow+1, end_column=7)
                ws1[f"A{xlrow}"]="Transporte"
                ws1[f"A{xlrow}"].fill = blueFill
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=14, bold=True)
                ws1[f"A{xlrow}"].border = medium_border
                xlrow+=3
                
                ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=4)
                ws1.merge_cells(start_row=xlrow, start_column=6, end_row=xlrow, end_column=7)
                ws1[f"A{xlrow}"]="Empresa"
                ws1[f"A{xlrow}"].fill = blueFill
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"A{xlrow}"].border = thin_border
                ws1[f"C{xlrow}"]="Conductor"
                ws1[f"C{xlrow}"].fill = blueFill
                ws1[f"C{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"C{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"C{xlrow}"].border = thin_border
                ws1[f"E{xlrow}"]="Teléfono"
                ws1[f"E{xlrow}"].fill = blueFill
                ws1[f"E{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"E{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"E{xlrow}"].border = thin_border
                ws1[f"F{xlrow}"]="Email"
                ws1[f"F{xlrow}"].fill = blueFill
                ws1[f"F{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"F{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"F{xlrow}"].border = thin_border
                xlrow+=1
                ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=4)
                ws1.merge_cells(start_row=xlrow, start_column=6, end_row=xlrow, end_column=7)
                ws1[f"A{xlrow}"]=data[0]
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=8)
                ws1[f"A{xlrow}"].border = thin_border
                ws1[f"C{xlrow}"]=data[1]
                ws1[f"C{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                ws1[f"C{xlrow}"].border = thin_border
                ws1[f"E{xlrow}"]=data[2]
                ws1[f"E{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"E{xlrow}"].font= Font(name='Arial',size=8)
                ws1[f"E{xlrow}"].border = thin_border
                ws1[f"F{xlrow}"]=data[3]
                ws1[f"F{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"F{xlrow}"].font= Font(name='Arial',size=8)
                ws1[f"F{xlrow}"].border = thin_border
                xlrow+=2
                
                ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=4)
                ws1.merge_cells(start_row=xlrow, start_column=5, end_row=xlrow, end_column=7)
                ws1[f"A{xlrow}"]="Vehículo"
                ws1[f"A{xlrow}"].fill = blueFill
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"A{xlrow}"].border = thin_border
                ws1[f"C{xlrow}"]="Matrícula"
                ws1[f"C{xlrow}"].fill = blueFill
                ws1[f"C{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"C{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"C{xlrow}"].border = thin_border
                ws1[f"E{xlrow}"]="Notas"
                ws1[f"E{xlrow}"].fill = blueFill
                ws1[f"E{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"E{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"E{xlrow}"].border = thin_border
                xlrow+=1
                ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=4)
                ws1.merge_cells(start_row=xlrow, start_column=5, end_row=xlrow, end_column=7)
                ws1[f"A{xlrow}"]=data[4]
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=8)
                ws1[f"A{xlrow}"].border = thin_border
                ws1[f"C{xlrow}"]=data[5]
                ws1[f"C{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                ws1[f"C{xlrow}"].border = thin_border
                ws1[f"E{xlrow}"]=data[6]
                ws1[f"E{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"E{xlrow}"].font= Font(name='Arial',size=8)
                ws1[f"E{xlrow}"].border = thin_border
                xlrow+=2
                
                #-------------------RECOGIDA-----------------------------------
                ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=7)
                ws1[f"A{xlrow}"]="Recogida"
                ws1[f"A{xlrow}"].fill = blueFill
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"A{xlrow}"].border = thin_border
                xlrow+=1
                
                ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                ws1[f"A{xlrow}"]="Lugar"
                ws1[f"A{xlrow}"].fill = blueFill
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"A{xlrow}"].border = thin_border
                ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                ws1[f"C{xlrow}"]=data[7]
                ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                ws1[f"C{xlrow}"].border = thin_border
                xlrow+=1
                
                ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                ws1[f"A{xlrow}"]="Localidad"
                ws1[f"A{xlrow}"].fill = blueFill
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"A{xlrow}"].border = thin_border
                ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                ws1[f"C{xlrow}"]=data[8]
                ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                ws1[f"C{xlrow}"].border = thin_border
                xlrow+=1
                
                ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                ws1[f"A{xlrow}"]="Dirección"
                ws1[f"A{xlrow}"].fill = blueFill
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"A{xlrow}"].border = thin_border
                ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                ws1[f"C{xlrow}"]=data[9]
                ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                ws1[f"C{xlrow}"].border = thin_border
                xlrow+=1
                
                ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                ws1[f"A{xlrow}"]="Indicaciones"
                ws1[f"A{xlrow}"].fill = blueFill
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"A{xlrow}"].border = thin_border
                ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                ws1[f"C{xlrow}"]=data[10]
                ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                ws1[f"C{xlrow}"].border = thin_border
                xlrow+=1
                
                ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                ws1[f"A{xlrow}"]="Fecha"
                ws1[f"A{xlrow}"].fill = blueFill
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"A{xlrow}"].border = thin_border
                ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                ws1[f"C{xlrow}"]=bdtrans.gira_fecha(data[11])
                ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                ws1[f"C{xlrow}"].border = thin_border
                xlrow+=1
                
                ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                ws1[f"A{xlrow}"]="Hora"
                ws1[f"A{xlrow}"].fill = blueFill
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"A{xlrow}"].border = thin_border
                ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                ws1[f"C{xlrow}"]=data[12]
                ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                ws1[f"C{xlrow}"].border = thin_border
                xlrow+=1
                
                #------------------ENTREGA------------------------------------
                ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=7)
                ws1[f"A{xlrow}"]="Entrega"
                ws1[f"A{xlrow}"].fill = blueFill
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"A{xlrow}"].border = thin_border
                xlrow+=1
                
                ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                ws1[f"A{xlrow}"]="Lugar"
                ws1[f"A{xlrow}"].fill = blueFill
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"A{xlrow}"].border = thin_border
                ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                ws1[f"C{xlrow}"]=data[13]
                ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                ws1[f"C{xlrow}"].border = thin_border
                xlrow+=1
                
                ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                ws1[f"A{xlrow}"]="Localidad"
                ws1[f"A{xlrow}"].fill = blueFill
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"A{xlrow}"].border = thin_border
                ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                ws1[f"C{xlrow}"]=data[14]
                ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                ws1[f"C{xlrow}"].border = thin_border
                xlrow+=1
                
                ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                ws1[f"A{xlrow}"]="Dirección"
                ws1[f"A{xlrow}"].fill = blueFill
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"A{xlrow}"].border = thin_border
                ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                ws1[f"C{xlrow}"]=data[15]
                ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                ws1[f"C{xlrow}"].border = thin_border
                xlrow+=1
                
                ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                ws1[f"A{xlrow}"]="Indicaciones"
                ws1[f"A{xlrow}"].fill = blueFill
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"A{xlrow}"].border = thin_border
                ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                ws1[f"C{xlrow}"]=data[16]
                ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                ws1[f"C{xlrow}"].border = thin_border
                xlrow+=1
                
                ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                ws1[f"A{xlrow}"]="Fecha"
                ws1[f"A{xlrow}"].fill = blueFill
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"A{xlrow}"].border = thin_border
                ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                ws1[f"C{xlrow}"]=bdtrans.gira_fecha(data[17])
                ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                ws1[f"C{xlrow}"].border = thin_border
                xlrow+=1
                
                ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                ws1[f"A{xlrow}"]="Hora"
                ws1[f"A{xlrow}"].fill = blueFill
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"A{xlrow}"].border = thin_border
                ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                ws1[f"C{xlrow}"]=data[18]
                ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                ws1[f"C{xlrow}"].border = thin_border
                xlrow+=2
                
                
                
                
            elif len(bdtrans.rows) > 1:
                trans = 1
                for data in bdtrans.rows:
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow+1, end_column=7)
                    ws1[f"A{xlrow}"]="Transporte"
                    ws1[f"A{xlrow}"].fill = blueFill
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=14, bold=True)
                    ws1[f"A{xlrow}"].border = medium_border
                    xlrow+=3
                    
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                    ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=4)
                    ws1.merge_cells(start_row=xlrow, start_column=6, end_row=xlrow, end_column=7)
                    ws1[f"A{xlrow}"]="Empresa"
                    ws1[f"A{xlrow}"].fill = blueFill
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"A{xlrow}"].border = thin_border
                    ws1[f"C{xlrow}"]="Conductor"
                    ws1[f"C{xlrow}"].fill = blueFill
                    ws1[f"C{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"C{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"C{xlrow}"].border = thin_border
                    ws1[f"E{xlrow}"]="Teléfono"
                    ws1[f"E{xlrow}"].fill = blueFill
                    ws1[f"E{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"E{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"E{xlrow}"].border = thin_border
                    ws1[f"F{xlrow}"]="Email"
                    ws1[f"F{xlrow}"].fill = blueFill
                    ws1[f"F{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"F{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"F{xlrow}"].border = thin_border
                    xlrow+=1
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                    ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=4)
                    ws1.merge_cells(start_row=xlrow, start_column=6, end_row=xlrow, end_column=7)
                    ws1[f"A{xlrow}"]=data[0]
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=8)
                    ws1[f"A{xlrow}"].border = thin_border
                    ws1[f"C{xlrow}"]=data[1]
                    ws1[f"C{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                    ws1[f"C{xlrow}"].border = thin_border
                    ws1[f"E{xlrow}"]=data[2]
                    ws1[f"E{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"E{xlrow}"].font= Font(name='Arial',size=8)
                    ws1[f"E{xlrow}"].border = thin_border
                    ws1[f"F{xlrow}"]=data[3]
                    ws1[f"F{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"F{xlrow}"].font= Font(name='Arial',size=8)
                    ws1[f"F{xlrow}"].border = thin_border
                    xlrow+=2
                    
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                    ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=4)
                    ws1.merge_cells(start_row=xlrow, start_column=5, end_row=xlrow, end_column=7)
                    ws1[f"A{xlrow}"]="Vehículo"
                    ws1[f"A{xlrow}"].fill = blueFill
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"A{xlrow}"].border = thin_border
                    ws1[f"C{xlrow}"]="Matrícula"
                    ws1[f"C{xlrow}"].fill = blueFill
                    ws1[f"C{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"C{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"C{xlrow}"].border = thin_border
                    ws1[f"E{xlrow}"]="Notas"
                    ws1[f"E{xlrow}"].fill = blueFill
                    ws1[f"E{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"E{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"E{xlrow}"].border = thin_border
                    xlrow+=1
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                    ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=4)
                    ws1.merge_cells(start_row=xlrow, start_column=5, end_row=xlrow, end_column=7)
                    ws1[f"A{xlrow}"]=data[4]
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=8)
                    ws1[f"A{xlrow}"].border = thin_border
                    ws1[f"C{xlrow}"]=data[5]
                    ws1[f"C{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                    ws1[f"C{xlrow}"].border = thin_border
                    ws1[f"E{xlrow}"]=data[6]
                    ws1[f"E{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"E{xlrow}"].font= Font(name='Arial',size=8)
                    ws1[f"E{xlrow}"].border = thin_border
                    xlrow+=2
                    
                    #-------------------RECOGIDA-----------------------------------
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=7)
                    ws1[f"A{xlrow}"]="Recogida"
                    ws1[f"A{xlrow}"].fill = blueFill
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"A{xlrow}"].border = thin_border
                    xlrow+=1
                    
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                    ws1[f"A{xlrow}"]="Lugar"
                    ws1[f"A{xlrow}"].fill = blueFill
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"A{xlrow}"].border = thin_border
                    ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                    ws1[f"C{xlrow}"]=data[7]
                    ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                    ws1[f"C{xlrow}"].border = thin_border
                    xlrow+=1
                    
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                    ws1[f"A{xlrow}"]="Localidad"
                    ws1[f"A{xlrow}"].fill = blueFill
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"A{xlrow}"].border = thin_border
                    ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                    ws1[f"C{xlrow}"]=data[8]
                    ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                    ws1[f"C{xlrow}"].border = thin_border
                    xlrow+=1
                    
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                    ws1[f"A{xlrow}"]="Dirección"
                    ws1[f"A{xlrow}"].fill = blueFill
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"A{xlrow}"].border = thin_border
                    ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                    ws1[f"C{xlrow}"]=data[9]
                    ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                    ws1[f"C{xlrow}"].border = thin_border
                    xlrow+=1
                    
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                    ws1[f"A{xlrow}"]="Indicaciones"
                    ws1[f"A{xlrow}"].fill = blueFill
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"A{xlrow}"].border = thin_border
                    ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                    ws1[f"C{xlrow}"]=data[10]
                    ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                    ws1[f"C{xlrow}"].border = thin_border
                    xlrow+=1
                    
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                    ws1[f"A{xlrow}"]="Fecha"
                    ws1[f"A{xlrow}"].fill = blueFill
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"A{xlrow}"].border = thin_border
                    ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                    ws1[f"C{xlrow}"]=bdtrans.gira_fecha(data[11])
                    ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                    ws1[f"C{xlrow}"].border = thin_border
                    xlrow+=1
                    
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                    ws1[f"A{xlrow}"]="Hora"
                    ws1[f"A{xlrow}"].fill = blueFill
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"A{xlrow}"].border = thin_border
                    ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                    ws1[f"C{xlrow}"]=data[12]
                    ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                    ws1[f"C{xlrow}"].border = thin_border
                    xlrow+=1
                    
                    #------------------ENTREGA------------------------------------
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=7)
                    ws1[f"A{xlrow}"]="Entrega"
                    ws1[f"A{xlrow}"].fill = blueFill
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"A{xlrow}"].border = thin_border
                    xlrow+=1
                    
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                    ws1[f"A{xlrow}"]="Lugar"
                    ws1[f"A{xlrow}"].fill = blueFill
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"A{xlrow}"].border = thin_border
                    ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                    ws1[f"C{xlrow}"]=data[13]
                    ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                    ws1[f"C{xlrow}"].border = thin_border
                    xlrow+=1
                    
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                    ws1[f"A{xlrow}"]="Localidad"
                    ws1[f"A{xlrow}"].fill = blueFill
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"A{xlrow}"].border = thin_border
                    ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                    ws1[f"C{xlrow}"]=data[14]
                    ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                    ws1[f"C{xlrow}"].border = thin_border
                    xlrow+=1
                    
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                    ws1[f"A{xlrow}"]="Dirección"
                    ws1[f"A{xlrow}"].fill = blueFill
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"A{xlrow}"].border = thin_border
                    ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                    ws1[f"C{xlrow}"]=data[15]
                    ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                    ws1[f"C{xlrow}"].border = thin_border
                    xlrow+=1
                    
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                    ws1[f"A{xlrow}"]="Indicaciones"
                    ws1[f"A{xlrow}"].fill = blueFill
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"A{xlrow}"].border = thin_border
                    ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                    ws1[f"C{xlrow}"]=data[16]
                    ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                    ws1[f"C{xlrow}"].border = thin_border
                    xlrow+=1
                    
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                    ws1[f"A{xlrow}"]="fecha"
                    ws1[f"A{xlrow}"].fill = blueFill
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"A{xlrow}"].border = thin_border
                    ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                    ws1[f"C{xlrow}"]=bdtrans.gira_fecha(data[17])
                    ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                    ws1[f"C{xlrow}"].border = thin_border
                    xlrow+=1
                    
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow, end_column=2)
                    ws1[f"A{xlrow}"]="Hora"
                    ws1[f"A{xlrow}"].fill = blueFill
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                    ws1[f"A{xlrow}"].border = thin_border
                    ws1.merge_cells(start_row=xlrow, start_column=3, end_row=xlrow, end_column=7)
                    ws1[f"C{xlrow}"]=data[18]
                    ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                    ws1[f"C{xlrow}"].border = thin_border
                    xlrow+=2
                xlrow+=1    
                
                
                
                #---------------cabecera del servicio-----------------------------------------
                
            bdserv=BdStd()
            bdserv.runsql(fr"""SELECT servicio FROM proveedores_evento WHERE id_evento= '{self.id_evento}' GROUP BY servicio;""")
            if len(bdserv.rows) > 0:
                
                ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow+1, end_column=7)
                ws1[f"A{xlrow}"]="Proveedores"
                ws1[f"A{xlrow}"].fill = magentaFill
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=14, bold=True)
                ws1[f"A{xlrow}"].border = medium_border
                
                for servicio in bdserv.rows:
                    
                    ws1.merge_cells(start_row=xlrow, start_column=1, end_row=xlrow+1, end_column=7)
                    ws1[f"A{xlrow}"]=servicio[0]
                    ws1[f"A{xlrow}"].fill = magentaFill
                    ws1[f"A{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws1[f"A{xlrow}"].font= Font(name='Arial',size=14, bold=True)
                    ws1[f"A{xlrow}"].border = medium_border
                    xlrow+=1
            #-----------cabecera de los proveedores con ese servicio-------------------
            
                
                xlrow+=2
                ws1.merge_cells(f"A{xlrow}:B{xlrow}")
                ws1[f"A{xlrow}"]="Contacto"
                ws1[f"A{xlrow}"].fill = magentaFill
                ws1[f"A{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"A{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"A{xlrow}"].border = medium_border
                
                ws1[f"C{xlrow}"]="Teléfono"
                ws1[f"C{xlrow}"].fill = magentaFill
                ws1[f"C{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"C{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"C{xlrow}"].border = medium_border
                
                ws1[f"D{xlrow}"]="Empresa"
                ws1[f"D{xlrow}"].fill = magentaFill
                ws1[f"D{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"D{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"D{xlrow}"].border = medium_border
                
                ws1.merge_cells(f"E{xlrow}:F{xlrow}")
                ws1[f"E{xlrow}"]="Notas"
                ws1[f"E{xlrow}"].fill = magentaFill
                ws1[f"E{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"E{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"E{xlrow}"].border = medium_border
                
                ws1[f"G{xlrow}"]="Fecha"
                ws1[f"G{xlrow}"].fill = magentaFill
                ws1[f"G{xlrow}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1[f"G{xlrow}"].font= Font(name='Arial',size=10, bold=True)
                ws1[f"G{xlrow}"].border = medium_border
                
                xlrow+=1
                
                bdprov=BdStd()
                bdprov.runsql(fr"""SELECT pe.servicio, contacto_onsite, telefono_onsite, pro.empresa, pe.notas,
                      dtrftime(%d-%m%Y,fecha)||" "||hora FROM proveedores_evento as pe, proveedores as pro 
                      WHERE pe.id_evento='{self.id_evento}'AND pe.id_proveedor=pro.id_proveedor AND pe.servicio='{servicio[0]}';""")
                print(servicio[0])
                if bdprov.rows != None:
                    for prov in bdprov.rows:
                        ws1.merge_cells(f"A{xlrow}:B{xlrow}")
                        ws1.merge_cells(f"E{xlrow}:F{xlrow}")
                        
                        ws1[f"A{xlrow}"]=prov[1]
                        ws1[f"A{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        ws1[f"A{xlrow}"].font= Font(name='Arial',size=8)
                        ws1[f"A{xlrow}"].border = thin_border
                        
                        ws1[f"C{xlrow}"]=prov[2]
                        ws1[f"C{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        ws1[f"C{xlrow}"].font= Font(name='Arial',size=8)
                        ws1[f"C{xlrow}"].border = thin_border
                        
                        ws1[f"D{xlrow}"]=prov[3]
                        ws1[f"D{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        ws1[f"D{xlrow}"].font= Font(name='Arial',size=8)
                        ws1[f"D{xlrow}"].border = thin_border
                        
                        ws1[f"E{xlrow}"]=prov[4]
                        ws1[f"E{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        ws1[f"E{xlrow}"].font= Font(name='Arial',size=8)
                        ws1[f"E{xlrow}"].border = thin_border
                        
                        ws1[f"G{xlrow}"]=prov[5]
                        ws1[f"G{xlrow}"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        ws1[f"G{xlrow}"].font= Font(name='Arial',size=8)
                        ws1[f"G{xlrow}"].border = thin_border
                        
                        xlrow+=2
   
            
    #-------------------------------Informa de la creción o del error----------
    
            fichero = self.carpeta + f"hojaderuta-{self.id_evento}.xlsx"
            wb.save(fichero)
            qm = QtWidgets.QMessageBox
            qm.information(self, 'Aleph', "Hoja de ruta creada correctamente")
            
        # except Exception as e:
        #     qm = QtWidgets.QMessageBox
        #     qm.warning(self, '', f"Error: {e}")
        
        
if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    ui = HojaDeRuta()
    ui.show()
    sys.exit(app.exec_())